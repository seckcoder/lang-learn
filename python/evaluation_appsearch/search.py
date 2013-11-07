#! /usr/bin/env python
#-*- coding:utf-8 -*
from openpyxl.reader.excel import load_workbook
import codecs
import sys

def get_type(i_kind):
	if (i_kind == u'应用名称'):
		return 'fine'
	elif (i_kind == u'泛需求'):
		return 'general'
	else:
		print 'Error' + i_kind
		sys.exit()
def print_score(d_score):
	print 'Total Score\t' + 'Total Num\t' + 'Bad num of first\t' + 'Bad num of first perc\t'
	for kind in d_score:
		item = d_score[kind]
		print kind, item[0] / item[1], item[1], item[2], float(item[2]) / item[1]
		print ""
	t_score = d_score['fine'][0] + d_score['general'][0]
	t_num = d_score['fine'][1] + d_score['general'][1]
	print 'Total', t_score / t_num

if __name__ == '__main__':
	wb = load_workbook(filename = r'evaluation_form.xlsx')
	debug_f = codecs.open('debug', 'w', 'gbk')
	sheetnames = wb.get_sheet_names()
	for name in sheetnames:
		print 'sheet ' + name
	ws = wb.get_sheet_by_name(sheetnames[2])
	print 'title:'+ws.title
	print 'sheet rows:' , ws.get_highest_row()
	print 'sheet cols:' , ws.get_highest_column()

	d_score = {
			'fine': [0]*3,  #tscore, tnum, badnum 1, badnum total
			'general':[0]*3
			}

	for rx in range(1, 149):
		w = [0]*6
		for cx in range(0,6):
			w[cx] = ws.cell(row = rx, column = cx).value
		kind = get_type(w[1])
		d_score[kind][0] += w[2]
		d_score[kind][1] += 1
		if (w[3] != 4):
			d_score[kind][2] += 1
		print w[0]
	print_score(d_score)
